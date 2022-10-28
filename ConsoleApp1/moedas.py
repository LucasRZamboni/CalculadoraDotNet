import pyautogui as py
import pandas as pd
import time
import json
import requests


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook

import pyperclip as clip
from datetime import date
import sys

# options = webdriver.ChromeOptions()
# options.add_argument("--headless")

acesso = requests.get('https://economia.awesomeapi.com.br/all')
cotacao = acesso.json()

Valor_dolar = round(float(cotacao['USD']['bid']),2)

tabela = load_workbook("Cotação.xlsx")
aba_ativa = tabela.active

for celula in aba_ativa["A"]:
    if celula.value == "Dolar":
        linha = celula.row
        aba_ativa[f"B{linha}"] = Valor_dolar
tabela.save("Cotação.xlsx")

Valor_euro = round(float(cotacao['EUR']['bid']),2)
tabela = load_workbook("Cotação.xlsx")
aba_ativa = tabela.active
for celula in aba_ativa["A"]:
    if celula.value == "Euro":
        linha = celula.row
        aba_ativa[f"B{linha}"] = Valor_euro
tabela.save("Cotação.xlsx")

Valor_peso = round(float(cotacao['ARS']['bid']),2)
tabela = load_workbook("Cotação.xlsx")
aba_ativa = tabela.active
for celula in aba_ativa["A"]:
    if celula.value == "Peso":
        linha = celula.row
        aba_ativa[f"B{linha}"] = Valor_peso
tabela.save("Cotação.xlsx")

Valor_yuan = round(float(cotacao['CNY']['bid']),2)
tabela = load_workbook("Cotação.xlsx")
aba_ativa = tabela.active
for celula in aba_ativa["A"]:
    if celula.value == "Yuan":
        linha = celula.row
        aba_ativa[f"B{linha}"] = Valor_yuan
tabela.save("Cotação.xlsx")





#navegador.quit()